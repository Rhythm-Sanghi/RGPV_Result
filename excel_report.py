import os
import time
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles.numbers import FORMAT_NUMBER_00

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
FAIL_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
FAIL_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
PASS_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
PASS_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
WARN_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
ALT_FILL  = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
NOREG_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
NOREG_FONT = Font(name="Calibri", italic=True, color="777777", size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
BORDER_SIDE = Side(style="thin", color="B8B8B8")
THIN_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _auto_width(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 3))


def _style_header_row(ws, row_num, last_col):
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _style_subheader_row(ws, row_num, last_col):
    for col in range(1, last_col + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def build_report(records: list[dict], output_dir: str, semester: str, course_type: str):
    # Safety: Ensure we have data to write
    if not records:
        return ""
    
    # Use all records, but filter out pure empty dictionaries
    valid_records = [r for r in records if r.get("roll_no")]
    if not valid_records:
        return ""

    os.makedirs(output_dir, exist_ok=True)
    xlsx_path = os.path.join(output_dir, "results.xlsx")

    print(f"  [OK] Updating results.xlsx with {len(valid_records)} students...")

    all_subjects = []
    for rec in valid_records:
        for sub in rec.get("subjects", {}).keys():
            if sub not in all_subjects:
                all_subjects.append(sub)
    all_subjects.sort()

    rows = []
    for rec in valid_records:
        row = {
            "Roll No": rec.get("roll_no", ""),
            "Name": rec.get("name", ""),
            "Result": rec.get("result_status", ""),
            "SGPA": _safe_float(rec.get("sgpa", "")),
            "CGPA": _safe_float(rec.get("cgpa", "")),
        }
        for sub in all_subjects:
            row[sub] = rec.get("subjects", {}).get(sub, "")
        rows.append(row)

    df = pd.DataFrame(rows)
    if not df.empty and "SGPA" in df.columns:
        df = df.sort_values("SGPA", ascending=False, na_position="last")
        df.insert(0, "Rank", range(1, len(df) + 1))

    wb = Workbook()
    _build_results_sheet(wb, df, all_subjects, semester, course_type)
    _build_analytics_sheet(wb, df, all_subjects, semester, course_type)
    _build_backlog_sheet(wb, df, all_subjects)

    # Robust save with retries (in case user has file open)
    max_retries = 3
    for attempt in range(max_retries):
        try:
            wb.save(xlsx_path)
            break
        except Exception as e:
            if attempt == max_retries - 1:
                raise e
            time.sleep(1)
    

    return xlsx_path


def _safe_float(val):
    try:
        return float(str(val).strip())
    except Exception:
        return None


def _build_results_sheet(wb: Workbook, df: pd.DataFrame, all_subjects: list, semester: str, course_type: str):
    ws = wb.active
    ws.title = "Results"
    ws.sheet_view.showGridLines = False

    title = f"RGPV {course_type} — Semester {semester} Result Analysis"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(6 + len(all_subjects) + 1, 8))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    base_headers = ["Rank", "Roll No", "Name", "Result", "SGPA", "CGPA"]
    headers = base_headers + all_subjects
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
    _style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    if df.empty:
        _auto_width(ws)
        return

    result_col_idx = headers.index("Result") + 1
    subject_start_col = len(base_headers) + 1

    for r_idx, (_, row) in enumerate(df.iterrows(), 3):
        is_alt = r_idx % 2 == 0
        result_text = str(row.get("Result", "") or "").strip().upper()
        is_not_registered = result_text == "NOT REGISTERED"

        for c_idx, header in enumerate(headers, 1):
            val = row.get(header, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=val)

            # NOT REGISTERED rows: grey + italic, skip all other colour logic
            if is_not_registered:
                cell.font = NOREG_FONT
                cell.fill = NOREG_FILL
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN if header in ("Rank", "Roll No", "SGPA", "CGPA") or c_idx >= subject_start_col else LEFT_ALIGN
                continue

            if row.get("Rank") in (1, 2, 3):
                cell.font = Font(name="Calibri", size=10, bold=True)
            else:
                cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if header in ("Rank", "SGPA", "CGPA"):
                cell.alignment = CENTER_ALIGN
            elif c_idx >= subject_start_col:
                cell.alignment = CENTER_ALIGN
                grade_val = str(val).strip().upper() if val else ""
                if grade_val == "F":
                    cell.fill = FAIL_FILL
                    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
                elif grade_val and grade_val != "":
                    cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            else:
                cell.alignment = LEFT_ALIGN

            if c_idx == result_col_idx:
                if "FAIL" in result_text or result_text == "F":
                    cell.fill = FAIL_FILL
                    cell.font = FAIL_FONT
                    cell.alignment = CENTER_ALIGN
                elif "PASS" in result_text or "CLEAR" in result_text:
                    cell.fill = PASS_FILL
                    cell.font = PASS_FONT
                    cell.alignment = CENTER_ALIGN
                elif is_alt:
                    cell.fill = ALT_FILL
            elif c_idx not in [result_col_idx] and c_idx < subject_start_col and is_alt:
                if not (c_idx >= subject_start_col and str(val).strip().upper() == "F"):
                    if not (c_idx >= subject_start_col and str(val).strip().upper() not in ("F", "")):
                        cell.fill = ALT_FILL

        ws.row_dimensions[r_idx].height = 16

    _auto_width(ws)


def _build_analytics_sheet(wb: Workbook, df: pd.DataFrame, all_subjects: list, semester: str, course_type: str):
    ws = wb.create_sheet("Analytics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    tc = ws.cell(row=1, column=1, value="📊  Batch Analytics Report")
    tc.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    tc.fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.cell(row=3, column=1, value="Summary").font = Font(bold=True, size=12, name="Calibri", color="1F3864")

    row = 4
    total = len(df)
    if total > 0 and "Result" in df.columns:
        pass_count = df["Result"].astype(str).str.upper().str.contains("PASS|CLEAR").sum()
        fail_count = total - pass_count
        pass_pct = round(pass_count / total * 100, 2) if total > 0 else 0
        avg_sgpa = round(df["SGPA"].dropna().mean(), 2) if "SGPA" in df.columns else "N/A"

        summary_data = [
            ("Total Students", total),
            ("Passed", pass_count),
            ("Failed", fail_count),
            ("Pass Percentage", f"{pass_pct}%"),
            ("Average SGPA", avg_sgpa),
        ]

        for label, value in summary_data:
            lc = ws.cell(row=row, column=1, value=label)
            vc = ws.cell(row=row, column=2, value=value)
            lc.font = Font(name="Calibri", bold=True, size=10, color="1F3864")
            vc.font = NORMAL_FONT
            lc.border = THIN_BORDER
            vc.border = THIN_BORDER
            lc.alignment = LEFT_ALIGN
            vc.alignment = CENTER_ALIGN
            lc.fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            row += 1

    row += 2
    ws.cell(row=row, column=1, value="Subject-wise Analysis").font = Font(bold=True, size=12, name="Calibri", color="1F3864")
    row += 1

    sub_headers = ["Subject Code", "Total Appeared", "Pass", "Fail", "Pass %", "Avg Grade Points"]
    for c, h in enumerate(sub_headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
    _style_header_row(ws, row, len(sub_headers))
    row += 1

    grade_points = {"O": 10, "A+": 9, "A": 8, "B+": 7, "B": 6, "C+": 5, "C": 4, "D+": 4, "D": 4, "F": 0, "EX": 10, "AB": 0}

    for sub in all_subjects:
        if sub not in df.columns:
            continue
        sub_series = df[sub].astype(str).str.strip().str.upper()
        appeared = sub_series[sub_series != ""].shape[0]
        failed = sub_series[sub_series == "F"].shape[0]
        passed_count = appeared - failed
        pass_pct_sub = round(passed_count / appeared * 100, 2) if appeared > 0 else 0
        points = [grade_points.get(g, None) for g in sub_series if g and g != "" and g != "NAN"]
        avg_pts = round(sum(p for p in points if p is not None) / len([p for p in points if p is not None]), 2) if points else ""

        cells_data = [sub, appeared, passed_count, failed, f"{pass_pct_sub}%", avg_pts]
        for c, val in enumerate(cells_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if c != 1 else LEFT_ALIGN
            if c == 4 and isinstance(val, (int, float)) and val > 0:
                cell.fill = FAIL_FILL
                cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        row += 1

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(40, max(10, max_len + 3))


def _build_backlog_sheet(wb: Workbook, df: pd.DataFrame, all_subjects: list):
    ws = wb.create_sheet("Backlog List")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    tc = ws.cell(row=1, column=1, value="⚠️  Students with Backlog / Fail Grade")
    tc.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    tc.fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    base_headers = ["Roll No", "Name", "Result", "SGPA", "CGPA", "Failed Subjects"]
    for c, h in enumerate(base_headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(base_headers))
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "A3"

    if df.empty:
        _auto_width(ws)
        return

    row = 3
    for _, rec in df.iterrows():
        result_text = str(rec.get("Result", "")).strip().upper()
        failed_subs = []
        for sub in all_subjects:
            if sub in rec and str(rec[sub]).strip().upper() == "F":
                failed_subs.append(sub)

        is_backlog = "FAIL" in result_text or result_text == "F" or len(failed_subs) > 0

        if not is_backlog:
            continue

        row_data = [
            rec.get("Roll No", ""),
            rec.get("Name", ""),
            rec.get("Result", ""),
            rec.get("SGPA", ""),
            rec.get("CGPA", ""),
            ", ".join(failed_subs),
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if c == 3:  # Result column (was 4, shifted left by removing Father's Name)
                cell.fill = FAIL_FILL
                cell.font = FAIL_FONT
                cell.alignment = CENTER_ALIGN
            elif c == 6:  # Failed Subjects column (was 7)
                cell.fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                cell.font = Font(name="Calibri", bold=True, color="8B0000", size=10)
                cell.alignment = LEFT_ALIGN
            else:
                cell.alignment = CENTER_ALIGN if c in (1, 4, 5) else LEFT_ALIGN
        row += 1

    _auto_width(ws)
